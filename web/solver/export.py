"""
Экспорт расписания в Excel-файл.

Формат: строки = дни месяца, столбцы = посты.
Дополнительный лист — сводка часов по сотрудникам.
"""

from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from data import Employee, MonthConfig, Post

DAY_NAMES_RU = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
WEEKEND_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
HOLIDAY_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")


def export_to_excel(
    solution: dict,
    posts: list[Post],
    employees: list[Employee],
    config: MonthConfig,
    filepath_or_buffer,
):
    wb = Workbook()
    _write_schedule_sheet(wb, solution, posts, config)
    _write_summary_sheet(wb, solution, employees, config)
    wb.save(filepath_or_buffer)
    if isinstance(filepath_or_buffer, str):
        print(f"Excel сохранён: {filepath_or_buffer}")


def _write_schedule_sheet(
    wb: Workbook,
    solution: dict,
    posts: list[Post],
    config: MonthConfig,
):
    ws = wb.active
    ws.title = "График"

    schedule = solution["schedule"]
    month_name = date(config.year, config.month, 1).strftime("%B %Y")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(posts) + 2)
    title_cell = ws.cell(row=1, column=1, value=f"ГРАФИК НА {month_name.upper()}  (N={config.norm_hours:.0f}ч.)")
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")

    # Pre-compute first shift after vacation
    first_after_vac: dict[str, int] = {}
    for emp_name, absent_days in config.absences.items():
        if not absent_days:
            continue
        last_absent = max(absent_days)
        for d in config.days:
            if d <= last_absent:
                continue
            for p in posts:
                if any(emp_name in person
                       for person in schedule.get(d, {}).get(p.id, [])):
                    first_after_vac[emp_name] = d
                    break
            if emp_name in first_after_vac:
                break

    headers = ["Дата", "Д/Н"] + [p.name for p in posts]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    for col_idx, post in enumerate(posts, 3):
        shift_label = f"({post.shift_hours}ч, {post.staff_required}чел)"
        cell = ws.cell(row=4, column=col_idx, value=shift_label)
        cell.font = Font(size=8, italic=True, color="666666")
        cell.alignment = Alignment(horizontal="center")

    row = 5
    for d in config.days:
        dt = date(config.year, config.month, d)
        dow_idx = dt.weekday()
        dow_name = DAY_NAMES_RU[dow_idx]
        is_wknd = config.is_weekend(d)
        is_hol = config.is_holiday(d)

        ws.cell(row=row, column=1, value=f"{d:02d}.{config.month:02d}").border = THIN_BORDER
        ws.cell(row=row, column=2, value=dow_name).border = THIN_BORDER

        for p_idx, post in enumerate(posts):
            col = p_idx + 3
            people = schedule.get(d, {}).get(post.id, [])
            if d not in config.post_active_days.get(post.id, []):
                val = "—"
            else:
                val = "\n".join(sorted(people)) if people else "?"

            bold = False
            for emp_name, fav_day in first_after_vac.items():
                if d == fav_day and emp_name in (people or []):
                    bold = True
                    break

            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = WRAP
            if bold:
                cell.font = Font(bold=True)

            if is_hol:
                cell.fill = HOLIDAY_FILL
            elif is_wknd:
                cell.fill = WEEKEND_FILL

        if is_hol:
            ws.cell(row=row, column=1).fill = HOLIDAY_FILL
            ws.cell(row=row, column=2).fill = HOLIDAY_FILL
        elif is_wknd:
            ws.cell(row=row, column=1).fill = WEEKEND_FILL
            ws.cell(row=row, column=2).fill = WEEKEND_FILL

        row += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 4
    for i, post in enumerate(posts):
        col_letter = get_column_letter(i + 3)
        ws.column_dimensions[col_letter].width = 18


def _write_summary_sheet(
    wb: Workbook,
    solution: dict,
    employees: list[Employee],
    config: MonthConfig,
):
    ws = wb.create_sheet("Сводка часов")

    headers = ["Сотрудник", "Ставка", "Отсутств.", "Цель (полн.)",
               "Цель (эфф.)", "Факт (ч)", "Δ"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    hours = solution.get("employee_hours", {})
    norm = config.norm_hours

    over_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    under_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

    for row_idx, emp in enumerate(sorted(employees, key=lambda e: e.name), 2):
        full_target = norm * emp.rate
        eff_target = config.employee_target_hours.get(emp.name, full_target)
        actual = hours.get(emp.name, 0)
        absent_count = len(config.absences.get(emp.name, []))
        delta = actual - eff_target

        ws.cell(row=row_idx, column=1, value=emp.name).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=emp.rate).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=absent_count or "").border = THIN_BORDER
        ws.cell(row=row_idx, column=4, value=round(full_target, 1)).border = THIN_BORDER
        ws.cell(row=row_idx, column=5,
                value=round(eff_target, 1) if absent_count else "").border = THIN_BORDER

        fact_cell = ws.cell(row=row_idx, column=6, value=actual)
        fact_cell.border = THIN_BORDER

        delta_cell = ws.cell(row=row_idx, column=7, value=round(delta, 1))
        delta_cell.border = THIN_BORDER
        if delta > 6:
            delta_cell.fill = over_fill
        elif delta < -6:
            delta_cell.fill = under_fill

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14
